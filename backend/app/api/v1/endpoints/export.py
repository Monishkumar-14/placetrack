import csv
import io
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy.orm import selectinload
from app.db.session import get_db
from app.deps import get_current_user
from app.models.user import User
from app.models.company import Company
from app.models.drive import PlacementDrive
from app.models.assessment import Assessment
from app.models.interview import Interview
from app.models.offer import Offer

router = APIRouter()


@router.get("/csv")
async def export_csv(db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Export all placement data as CSV"""
    result = await db.execute(
        select(Company)
        .where(Company.user_id == current_user.id)
        .options(
            selectinload(Company.drives)
            .selectinload(PlacementDrive.assessments),
            selectinload(Company.drives)
            .selectinload(PlacementDrive.interviews),
            selectinload(Company.drives)
            .selectinload(PlacementDrive.offer),
        )
        .order_by(Company.name)
    )
    companies = result.scalars().unique().all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Company", "Industry", "Priority", "Job Role", "Job Location",
        "Drive Type", "Application Date", "Status", "Current Stage",
        "Assessments Count", "Interviews Count", "Has Offer",
        "Offer Type", "Monthly Stipend", "Annual CTC", "PPO Available",
        "PPO CTC", "Offer Status"
    ])

    for company in companies:
        if not company.drives:
            writer.writerow([
                company.name, company.industry or "", company.priority or "",
                "", "", "", "", "", "", 0, 0, "No",
                "", "", "", "", "", ""
            ])
        for drive in company.drives:
            assessments_count = len(drive.assessments) if drive.assessments else 0
            interviews_count = len(drive.interviews) if drive.interviews else 0
            offer = drive.offer

            writer.writerow([
                company.name,
                company.industry or "",
                company.priority or "",
                drive.job_role or "",
                drive.job_location or "",
                drive.drive_type or "",
                str(drive.application_date) if drive.application_date else "",
                drive.overall_status or "",
                drive.current_stage or "",
                assessments_count,
                interviews_count,
                "Yes" if offer else "No",
                offer.offer_type if offer else "",
                str(offer.monthly_stipend) if offer and offer.monthly_stipend else "",
                str(offer.annual_ctc) if offer and offer.annual_ctc else "",
                "Yes" if offer and offer.ppo_available else "No" if offer else "",
                str(offer.ppo_ctc) if offer and offer.ppo_ctc else "",
                offer.offer_status if offer else "",
            ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=placement_data.csv"}
    )


@router.get("/excel")
async def export_excel(db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Export all placement data as Excel"""
    try:
        from openpyxl import Workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    result = await db.execute(
        select(Company)
        .where(Company.user_id == current_user.id)
        .options(
            selectinload(Company.drives)
            .selectinload(PlacementDrive.assessments),
            selectinload(Company.drives)
            .selectinload(PlacementDrive.interviews),
            selectinload(Company.drives)
            .selectinload(PlacementDrive.offer),
        )
        .order_by(Company.name)
    )
    companies = result.scalars().unique().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Placement Data"

    headers = [
        "Company", "Industry", "Priority", "Job Role", "Job Location",
        "Drive Type", "Application Date", "Status", "Current Stage",
        "Assessments", "Interviews", "Offer Type", "Stipend",
        "Annual CTC", "PPO Available", "PPO CTC", "Offer Status"
    ]
    ws.append(headers)

    # Style header
    from openpyxl.styles import Font, PatternFill
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")

    for company in companies:
        for drive in company.drives:
            offer = drive.offer
            ws.append([
                company.name,
                company.industry or "",
                company.priority or "",
                drive.job_role or "",
                drive.job_location or "",
                drive.drive_type or "",
                str(drive.application_date) if drive.application_date else "",
                drive.overall_status or "",
                drive.current_stage or "",
                len(drive.assessments) if drive.assessments else 0,
                len(drive.interviews) if drive.interviews else 0,
                offer.offer_type if offer else "",
                offer.monthly_stipend if offer and offer.monthly_stipend else "",
                offer.annual_ctc if offer and offer.annual_ctc else "",
                "Yes" if offer and offer.ppo_available else "",
                offer.ppo_ctc if offer and offer.ppo_ctc else "",
                offer.offer_status if offer else "",
            ])

    # Auto-size columns
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 30)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=placement_data.xlsx"}
    )


@router.post("/import/csv")
async def import_csv(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Import placement data from CSV"""
    if not file.filename or not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="Please upload a CSV file")

    content = await file.read()
    text = content.decode('utf-8')
    reader = csv.DictReader(io.StringIO(text))

    imported = 0
    for row in reader:
        company_name = row.get('Company', '').strip()
        if not company_name:
            continue

        # Check if company exists
        existing = await db.execute(
            select(Company).where(Company.user_id == current_user.id, Company.name == company_name)
        )
        company = existing.scalar_one_or_none()

        if not company:
            company = Company(
                user_id=current_user.id,
                name=company_name,
                industry=row.get('Industry', '').strip() or None,
                priority=row.get('Priority', 'none').strip() or 'none',
            )
            db.add(company)
            await db.flush()

        # Create drive if role specified
        job_role = row.get('Job Role', row.get('Role', '')).strip()
        if job_role:
            drive = PlacementDrive(
                company_id=company.id,
                job_role=job_role,
                job_location=row.get('Job Location', '').strip() or None,
                drive_type=row.get('Drive Type', 'campus').strip() or 'campus',
                overall_status=row.get('Status', 'applied').strip() or 'applied',
            )
            db.add(drive)

        imported += 1

    await db.commit()
    return {"imported": imported, "message": f"Successfully imported {imported} records"}
